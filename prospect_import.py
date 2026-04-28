"""
Parse CSV / Excel prospect lists into normalized dict rows.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "first_name": ("first name", "firstname", "first", "given name", "fname"),
    "last_name": ("last name", "lastname", "last", "lname", "surname"),
    "phone": (
        "phone",
        "mobile",
        "number",
        "phone number",
        "telephone",
        "tel",
        "cell",
        "phone_number",
        "direct phone",
        "direct_phone",
        "work phone",
        "work_phone",
        "contact number",
        "contact_number",
        "mobile phone",
        "mobile_phone",
        "corporate phone",
        "person direct phone",
        "# phone",
    ),
    "title": ("title", "job title", "role", "position", "job_title", "person title", "designation"),
    "company": ("company", "organization", "org", "account", "employer", "company name", "organization name", "company_name"),
    "notes": ("notes", "context", "description", "comments", "note"),
    "email": ("email", "e-mail", "work email", "email address", "work_email", "person email"),
}


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def _map_headers(row_keys: list[str]) -> dict[str, str]:
    """Map normalized header -> canonical field name."""
    inv: dict[str, str] = {}
    for canon, aliases in HEADER_ALIASES.items():
        for a in aliases:
            inv[_norm_header(a)] = canon
    out: dict[str, str] = {}
    for k in row_keys:
        nk = _norm_header(k)
        if nk in inv:
            out[k] = inv[nk]
    return out


def _rows_from_dicts(dict_rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    if not dict_rows:
        return []
    keys = list(dict_rows[0].keys())
    hmap = _map_headers(keys)
    out: list[dict[str, str]] = []
    for raw in dict_rows:
        row = {
            "first_name": "",
            "last_name": "",
            "phone": "",
            "title": "",
            "company": "",
            "notes": "",
            "email": "",
            "apollo_person_id": "",
        }
        for src_key, canon in hmap.items():
            val = raw.get(src_key)
            if val is not None:
                row[canon] = str(val).strip()
        # pass through extra columns into notes if unmapped
        out.append(row)
    return out


def parse_csv_bytes(data: bytes) -> tuple[list[dict[str, str]], list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], ["No header row found"]
    dict_rows = [r for r in reader if any((v or "").strip() for v in r.values())]
    rows = _rows_from_dicts(dict_rows)
    warnings: list[str] = []
    if rows and not any(r.get("phone") for r in rows):
        warnings.append("No phone column detected — use a column named Phone, Mobile, or Number.")
    return rows, warnings


def parse_xlsx_bytes(data: bytes) -> tuple[list[dict[str, str]], list[str]]:
    from openpyxl import load_workbook

    bio = io.BytesIO(data)
    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["Empty spreadsheet"]
    headers = [str(c) if c is not None else "" for c in header]
    dict_rows: list[dict[str, Any]] = []
    for tup in rows_iter:
        if not tup or all(v is None or str(v).strip() == "" for v in tup):
            continue
        d = {}
        for i, h in enumerate(headers):
            if i < len(tup) and tup[i] is not None:
                d[h] = tup[i]
            else:
                d[h] = ""
        dict_rows.append(d)
    rows = _rows_from_dicts(dict_rows)
    warnings: list[str] = []
    if rows and not any(r.get("phone") for r in rows):
        warnings.append("No phone column detected — use a column named Phone, Mobile, or Number.")
    return rows, warnings
